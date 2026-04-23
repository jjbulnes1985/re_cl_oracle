"""
export_scraped.py
-----------------
Exporta los listings scrapeados (Portal Inmobiliario + Toctoc) con sus scores
a CSV y Excel para uso en planilla.

Requiere que scraped_to_scored.py haya corrido previamente para tener scores.
Si no hay scores, igual exporta los listings crudos con columnas de score vacías.

Salida:
  data/exports/scraped_listings_YYYY-MM-DD.csv   (siempre)
  data/exports/scraped_listings_YYYY-MM-DD.xlsx  (si openpyxl disponible)

Uso:
    py scripts/export_scraped.py
    py scripts/export_scraped.py --min-score 0.6
    py scripts/export_scraped.py --source portal_inmobiliario
    py scripts/export_scraped.py --type apartments
    py scripts/export_scraped.py --since 2024-06-01
    py scripts/export_scraped.py --top-n 500
    py scripts/export_scraped.py --no-excel      # solo CSV
"""

import argparse
import os
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from loguru import logger
from sqlalchemy import create_engine, text

load_dotenv()
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))


# ── DB connection ─────────────────────────────────────────────────────────────

def _build_db_url() -> str:
    url = os.getenv("DATABASE_URL")
    if url:
        return url
    host = os.getenv("POSTGRES_HOST", "localhost")
    port = os.getenv("POSTGRES_PORT", "5432")
    db   = os.getenv("POSTGRES_DB",   "re_cl")
    user = os.getenv("POSTGRES_USER", "re_cl_user")
    pwd  = os.getenv("POSTGRES_PASSWORD", "")
    return f"postgresql://{user}:{pwd}@{host}:{port}/{db}"


# ── Query ─────────────────────────────────────────────────────────────────────

QUERY = """
    SELECT
        sl.id                                           AS id_interno,
        sl.source                                       AS fuente,
        sl.external_id                                  AS id_portal,
        sl.project_type                                 AS tipo,
        sl.county_name                                  AS comuna,
        sl.address                                      AS direccion,
        sl.price_uf                                     AS precio_uf,
        sl.surface_m2                                   AS superficie_m2,
        sl.uf_m2                                        AS uf_por_m2_publicado,
        sl.bedrooms                                     AS dormitorios,
        sl.bathrooms                                    AS banos,
        sl.latitude                                     AS latitud,
        sl.longitude                                    AS longitud,
        sl.scraped_at                                   AS fecha_scraping,
        sl.url                                          AS url,
        -- Scores (NULL si no se ha corrido scraped_to_scored.py)
        ms.opportunity_score                            AS score_oportunidad,
        ms.undervaluation_score                         AS score_subvaloracion,
        ms.gap_pct                                      AS gap_pct,
        ms.predicted_uf_m2                              AS uf_m2_precio_justo,
        ms.data_confidence                              AS confianza_dato,
        ms.scoring_profile                              AS perfil_scoring,
        ms.scored_at                                    AS fecha_scoring,
        -- Gap en % legible
        ROUND(CAST(ms.gap_pct * 100 AS numeric), 1)    AS gap_porcentaje,
        -- Diferencia de precio total
        ROUND(CAST(
            (ms.gap_pct * sl.surface_m2 * ms.predicted_uf_m2) AS numeric
        ), 0)                                           AS diferencia_uf_total
    FROM scraped_listings sl
    LEFT JOIN model_scores ms
        ON ms.clean_id = sl.id
        AND ms.source = 'scraped'
    WHERE 1=1
    {filters}
    ORDER BY ms.opportunity_score DESC NULLS LAST, sl.scraped_at DESC
    {limit_clause}
"""


def load_data(
    engine,
    min_score: float = 0.0,
    source: str = None,
    project_type: str = None,
    since: str = None,
    top_n: int = None,
) -> pd.DataFrame:
    filter_parts = []
    if min_score > 0:
        filter_parts.append(f"AND ms.opportunity_score >= {min_score}")
    if source:
        filter_parts.append(f"AND sl.source = '{source}'")
    if project_type:
        filter_parts.append(f"AND sl.project_type = '{project_type}'")
    if since:
        filter_parts.append(f"AND sl.scraped_at >= '{since}'")

    limit_clause = f"LIMIT {top_n}" if top_n else ""
    filters_str  = "\n    ".join(filter_parts)

    query = QUERY.format(filters=filters_str, limit_clause=limit_clause)
    df = pd.read_sql(text(query), engine)
    logger.info(f"Cargados {len(df):,} listings desde la DB")
    return df


# ── Formato planilla ──────────────────────────────────────────────────────────

COLUMN_ORDER = [
    "fuente", "tipo", "comuna", "precio_uf", "superficie_m2",
    "uf_por_m2_publicado", "uf_m2_precio_justo", "gap_porcentaje",
    "diferencia_uf_total", "score_oportunidad", "score_subvaloracion",
    "confianza_dato", "dormitorios", "banos", "direccion",
    "latitud", "longitud", "url", "fecha_scraping",
    "perfil_scoring", "id_portal", "fuente", "id_interno",
]

# Solo columnas que existen en el df
def ordered_columns(df: pd.DataFrame) -> list:
    seen = set()
    result = []
    for col in COLUMN_ORDER:
        if col in df.columns and col not in seen:
            result.append(col)
            seen.add(col)
    # Agregar cualquier columna restante
    for col in df.columns:
        if col not in seen:
            result.append(col)
    return result


def format_df(df: pd.DataFrame) -> pd.DataFrame:
    """Reordena columnas y formatea tipos para planilla."""
    df = df[ordered_columns(df)].copy()

    # Redondear flotantes
    for col in ["precio_uf", "uf_por_m2_publicado", "uf_m2_precio_justo"]:
        if col in df.columns:
            df[col] = df[col].round(1)
    for col in ["score_oportunidad", "score_subvaloracion", "confianza_dato"]:
        if col in df.columns:
            df[col] = df[col].round(3)

    # Fecha sin timezone para Excel
    for col in ["fecha_scraping", "fecha_scoring"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col]).dt.tz_localize(None)

    return df


# ── Export ────────────────────────────────────────────────────────────────────

def export(df: pd.DataFrame, exports_dir: Path, excel: bool = True) -> None:
    today = date.today().isoformat()
    exports_dir.mkdir(parents=True, exist_ok=True)

    # CSV (siempre)
    csv_path = exports_dir / f"scraped_listings_{today}.csv"
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    logger.info(f"CSV exportado: {csv_path}  ({len(df):,} filas)")

    # Excel
    if not excel:
        return
    try:
        import openpyxl  # noqa: F401
        xlsx_path = exports_dir / f"scraped_listings_{today}.xlsx"
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Listings")

            # Dar formato al header
            ws = writer.sheets["Listings"]
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Autowidth (estimado)
            for col in ws.columns:
                max_len = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in col
                )
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

            # Freeze header row
            ws.freeze_panes = "A2"

        logger.info(f"Excel exportado: {xlsx_path}  ({len(df):,} filas)")
    except ImportError:
        logger.warning("openpyxl no instalado — solo CSV. Instalar con: pip install openpyxl")


# ── Summary ───────────────────────────────────────────────────────────────────

def print_summary(df: pd.DataFrame) -> None:
    logger.info("=" * 60)
    logger.info(f"TOTAL listings: {len(df):,}")

    if "fuente" in df.columns:
        logger.info("\nPor fuente:")
        for fuente, n in df["fuente"].value_counts().items():
            logger.info(f"  {fuente:<30} {n:>6,}")

    if "tipo" in df.columns:
        logger.info("\nPor tipo:")
        for tipo, n in df["tipo"].value_counts().items():
            logger.info(f"  {tipo:<30} {n:>6,}")

    if "comuna" in df.columns:
        logger.info("\nTop 10 comunas:")
        for comuna, n in df["comuna"].value_counts().head(10).items():
            logger.info(f"  {comuna:<30} {n:>6,}")

    if "score_oportunidad" in df.columns and df["score_oportunidad"].notna().any():
        scored = df[df["score_oportunidad"].notna()]
        high   = (scored["score_oportunidad"] >= 0.7).sum()
        logger.info(f"\nListings con score:    {len(scored):,}")
        logger.info(f"Score alto (>=0.70):   {high:,}")
        logger.info(f"Score promedio:        {scored['score_oportunidad'].mean():.3f}")

        if "gap_porcentaje" in df.columns:
            top5 = scored.nlargest(5, "score_oportunidad")[
                ["comuna", "tipo", "precio_uf", "uf_por_m2_publicado",
                 "uf_m2_precio_justo", "gap_porcentaje", "score_oportunidad", "url"]
            ]
            logger.info("\nTOP 5 OPORTUNIDADES:")
            for _, r in top5.iterrows():
                logger.info(
                    f"  {str(r.get('comuna','')):<20} {str(r.get('tipo','')):<12} "
                    f"UF {r.get('precio_uf',''):<8}  "
                    f"gap={r.get('gap_porcentaje',''):.1f}%  "
                    f"score={r.get('score_oportunidad',''):.3f}"
                )
    else:
        logger.warning("Sin scores — corre primero: py src/scoring/scraped_to_scored.py")

    logger.info("=" * 60)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Exportar listings scrapeados a CSV/Excel")
    parser.add_argument("--min-score",  type=float, default=0.0,
                        help="Score mínimo para incluir (0-1). Default: todos.")
    parser.add_argument("--source",     type=str,   default=None,
                        choices=["portal_inmobiliario", "toctoc"],
                        help="Filtrar por fuente")
    parser.add_argument("--type",       type=str,   default=None,
                        choices=["apartments", "residential", "land", "retail"],
                        help="Filtrar por tipo de propiedad")
    parser.add_argument("--since",      type=str,   default=None,
                        help="Solo listings desde esta fecha (YYYY-MM-DD)")
    parser.add_argument("--top-n",      type=int,   default=None,
                        help="Máximo de filas a exportar (ordenado por score)")
    parser.add_argument("--no-excel",   action="store_true",
                        help="Exportar solo CSV, sin Excel")
    parser.add_argument("--output-dir", type=str,
                        default=str(Path(__file__).resolve().parents[1] / "data" / "exports"),
                        help="Directorio de salida")
    args = parser.parse_args()

    engine      = create_engine(_build_db_url(), pool_pre_ping=True)
    exports_dir = Path(args.output_dir)

    df = load_data(
        engine,
        min_score    = args.min_score,
        source       = args.source,
        project_type = args.type,
        since        = args.since,
        top_n        = args.top_n,
    )

    if df.empty:
        logger.warning("No hay listings que exportar. ¿Corriste los scrapers?")
        logger.warning("  py src/scraping/portal_inmobiliario.py --max-pages 50")
        logger.warning("  py src/scraping/toctoc.py --max-pages 50")
        return

    print_summary(df)
    df = format_df(df)
    export(df, exports_dir, excel=not args.no_excel)


if __name__ == "__main__":
    main()
